"""
MLB Player Props Spreadsheet — EXPERIMENTAL.
============================================
Sandbox exporter. Builds a daily spreadsheet of player-prop projections
for offline auditing only. Outputs land in `data/experimental/mlb-props/`,
which is deliberately OUTSIDE `public/` so the website can't surface
this data and it never appears in the daily card.

Per BRAND_GUIDE Sandbox protocol: prop markets do not ship to the
daily card or anywhere user-facing until they pass the same gate as
game-level markets (≥+1% ROI AND Brier <0.246 over 200+ bets, derived
from prop-specific backtest).

Outputs per run:
    data/experimental/mlb-props/
        mlb_props_<date>.json     — structured payload for inspection
        mlb_props_<date>.xlsx     — multi-tab workbook for human eyeballs
        mlb_props_<date>.csv      — flat per-projection rows for filtering

Tabs / sections:
    Pitcher Ks       — every probable SP × {O4.5, O5.5, O6.5, O7.5}
    Batter Hits      — every starter in posted lineup × {O0.5, O1.5}
    Batter Total Bases — every starter × {O1.5, O2.5, O3.5}

Cron-callable. Idempotent — overwrites today's outputs in place.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Iterable

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None

from scrapers.mlb.mlb_game_scraper import MLBGameScraper
from scrapers.mlb.mlb_pitcher_scraper import MLBPitcherScraper
from scrapers.mlb.mlb_player_props_scraper import MLBPlayerPropsScraper
from exporters.mlb.daily_spreadsheet import fetch_slate
from models.mlb.player_props import (
    pitcher_strikeouts,
    batter_hits,
    batter_total_bases,
    avg_ip_per_start,
    expected_abs_for_lineup_slot,
    PITCHER_K_LINES,
    BATTER_HITS_LINES,
    BATTER_TB_LINES,
)

DEFAULT_OUTPUT_DIR = REPO_ROOT / "data" / "experimental" / "mlb-props"
SEASON_DEFAULT = 2026


def _today_et() -> str:
    if ZoneInfo is not None:
        return datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d")
    return datetime.utcnow().strftime("%Y-%m-%d")


class PlayerPropsExperimental:
    """Builds and writes the daily player-props sandbox spreadsheet."""

    def __init__(
        self,
        season: int = SEASON_DEFAULT,
        target_date: str | None = None,
        output_dir: Path | None = None,
    ):
        self.season = season
        self.target_date = target_date or _today_et()
        self.output_dir = Path(output_dir) if output_dir else DEFAULT_OUTPUT_DIR
        self.pitcher_scraper = MLBPitcherScraper(season=season)
        self.props_scraper = MLBPlayerPropsScraper(season=season)
        self.props_scraper._lineup_scraper.season = season  # ensure aligned

    # ---------------- collect ---------------------------------------

    def collect(self) -> dict:
        print(f"  Fetching slate for {self.target_date}...")
        slate = fetch_slate(self.target_date)
        print(f"    {len(slate)} scheduled games")
        if not slate:
            return self._empty_payload()

        print("  Fetching probable starting pitchers...")
        sp_map = self.pitcher_scraper.fetch_factors_for_slate(slate)

        print("  Fetching player-prop data (lineups + season stats)...")
        slate_data = self.props_scraper.fetch_for_slate(slate, sp_map=sp_map)

        # Project each prop type from the assembled data.
        pitcher_ks_rows: list[dict] = []
        batter_hits_rows: list[dict] = []
        batter_tb_rows: list[dict] = []

        for game_pk, sides in slate_data.items():
            away = sides["away_team"]
            home = sides["home_team"]
            matchup = f"{away}@{home}"

            for side_label, side, opp_side in (
                ("AWAY", sides["away"], sides["home"]),
                ("HOME", sides["home"], sides["away"]),
            ):
                pitcher_ks_rows.extend(
                    self._project_pitcher_ks(matchup, side_label, side, opp_side)
                )
                batter_hits_rows.extend(
                    self._project_batter_hits(matchup, side_label, side, opp_side)
                )
                batter_tb_rows.extend(
                    self._project_batter_tb(matchup, side_label, side, opp_side)
                )

        return {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "target_date": self.target_date,
            "season": self.season,
            "experimental": True,
            "warning": (
                "EXPERIMENTAL — projections are sandboxed for offline "
                "auditing only. Not on the daily card; not on the website. "
                "Per BRAND_GUIDE Sandbox protocol, no prop market ships "
                "until it passes the same gate as game-level markets."
            ),
            "counts": {
                "slate_games": len(slate),
                "pitcher_k_rows": len(pitcher_ks_rows),
                "batter_hits_rows": len(batter_hits_rows),
                "batter_tb_rows": len(batter_tb_rows),
            },
            "pitcher_strikeouts": pitcher_ks_rows,
            "batter_hits": batter_hits_rows,
            "batter_total_bases": batter_tb_rows,
        }

    # ---------------- per-prop projections ------------------------------

    @staticmethod
    def _project_pitcher_ks(
        matchup: str, side_label: str, side: dict, opp_side: dict,
    ) -> list[dict]:
        sp_id = side.get("pitcher_id")
        sp_name = side.get("pitcher_name")
        sp_stats = side.get("pitcher_stats")
        if not sp_id or not sp_stats:
            return []

        opp_k = opp_side.get("team_k_per_9")
        ip_estimate = avg_ip_per_start(sp_stats.get("ip"), sp_stats.get("starts"))
        proj = pitcher_strikeouts(
            season_ks=sp_stats.get("ks"),
            season_ip=sp_stats.get("ip"),
            opp_team_k_per_9=opp_k,
            expected_ip_today=ip_estimate,
        )

        return [{
            "matchup": matchup,
            "side": side_label,
            "pitcher_id": sp_id,
            "pitcher": sp_name,
            "season_ks": sp_stats.get("ks"),
            "season_ip": sp_stats.get("ip"),
            "season_baa": sp_stats.get("baa"),
            "opp_team_k_per_9": (
                round(opp_k, 2) if opp_k is not None else None
            ),
            **proj,
        }]

    @staticmethod
    def _project_batter_hits(
        matchup: str, side_label: str, side: dict, opp_side: dict,
    ) -> list[dict]:
        opp_pitcher_baa = (opp_side.get("pitcher_stats") or {}).get("baa")
        rows: list[dict] = []
        for slot, batter_id in enumerate(side.get("batter_ids") or [], 1):
            stats = PlayerPropsExperimental._fetch_or_none(side, batter_id)
            if stats is None:
                continue
            proj = batter_hits(
                season_avg=stats.get("avg"),
                season_ab=stats.get("ab"),
                expected_abs=expected_abs_for_lineup_slot(slot),
                opp_pitcher_baa=opp_pitcher_baa,
            )
            rows.append({
                "matchup": matchup,
                "side": side_label,
                "lineup_slot": slot,
                "batter_id": batter_id,
                "season_avg": stats.get("avg"),
                "season_ab": stats.get("ab"),
                "opp_pitcher_baa": opp_pitcher_baa,
                **proj,
            })
        return rows

    @staticmethod
    def _project_batter_tb(
        matchup: str, side_label: str, side: dict, opp_side: dict,
    ) -> list[dict]:
        opp_pitcher_baa = (opp_side.get("pitcher_stats") or {}).get("baa")
        rows: list[dict] = []
        for slot, batter_id in enumerate(side.get("batter_ids") or [], 1):
            stats = PlayerPropsExperimental._fetch_or_none(side, batter_id)
            if stats is None:
                continue
            proj = batter_total_bases(
                season_slg=stats.get("slg"),
                season_ab=stats.get("ab"),
                expected_abs=expected_abs_for_lineup_slot(slot),
                opp_pitcher_baa=opp_pitcher_baa,
            )
            rows.append({
                "matchup": matchup,
                "side": side_label,
                "lineup_slot": slot,
                "batter_id": batter_id,
                "season_slg": stats.get("slg"),
                "season_ab": stats.get("ab"),
                "opp_pitcher_baa": opp_pitcher_baa,
                **proj,
            })
        return rows

    @staticmethod
    def _fetch_or_none(side: dict, batter_id: int) -> dict | None:
        # Side cache: lazily attach a stats dict per side so we only hit
        # the API once per batter regardless of how many prop types use it.
        cache = side.setdefault("_batter_stats_cache", {})
        if batter_id in cache:
            return cache[batter_id]
        # The props_scraper holds the network call; we look up via that.
        # Find the ambient scraper instance through closure isn't possible
        # in a staticmethod — caller wires it via a different path. For
        # MVP we allow re-fetch: cache below de-dupes within a side.
        from scrapers.mlb.mlb_player_props_scraper import MLBPlayerPropsScraper
        # Reuse cached batter stats from a module-level scraper if one
        # was created earlier in this run (cheap lookup, no new fetches
        # since the same player_id is cached on the scraper instance).
        scraper = _MODULE_SCRAPER
        stats = scraper.fetch_batter_stats(batter_id) if scraper else None
        cache[batter_id] = stats
        return stats

    @staticmethod
    def _empty_payload() -> dict:
        return {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "target_date": None,
            "season": None,
            "experimental": True,
            "warning": "No games on slate.",
            "counts": {"slate_games": 0},
            "pitcher_strikeouts": [],
            "batter_hits": [],
            "batter_total_bases": [],
        }

    # ---------------- write -----------------------------------------

    def write(self, payload: dict) -> list[Path]:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        date = payload.get("target_date") or self.target_date
        written: list[Path] = []

        json_path = self.output_dir / f"mlb_props_{date}.json"
        json_path.write_text(json.dumps(payload, indent=2, default=str))
        written.append(json_path)

        csv_path = self.output_dir / f"mlb_props_{date}.csv"
        self._write_csv(csv_path, payload)
        written.append(csv_path)

        xlsx_path = self.output_dir / f"mlb_props_{date}.xlsx"
        self._write_xlsx(xlsx_path, payload)
        written.append(xlsx_path)

        return written

    @staticmethod
    def _write_csv(path: Path, payload: dict) -> None:
        rows: list[dict] = []
        for r in payload.get("pitcher_strikeouts", []):
            for line in PITCHER_K_LINES:
                key = f"over_{line}".replace(".", "_")
                rows.append({
                    "market": "pitcher_ks",
                    "matchup": r["matchup"],
                    "side": r["side"],
                    "player": r["pitcher"],
                    "line": line,
                    "expected": r.get("expected_ks"),
                    "model_prob_over": r.get(key),
                })
        for r in payload.get("batter_hits", []):
            for line in BATTER_HITS_LINES:
                key = f"over_{line}".replace(".", "_")
                rows.append({
                    "market": "batter_hits",
                    "matchup": r["matchup"],
                    "side": r["side"],
                    "player": f"slot {r['lineup_slot']} (id {r['batter_id']})",
                    "line": line,
                    "expected": r.get("expected_hits"),
                    "model_prob_over": r.get(key),
                })
        for r in payload.get("batter_total_bases", []):
            for line in BATTER_TB_LINES:
                key = f"over_{line}".replace(".", "_")
                rows.append({
                    "market": "batter_tb",
                    "matchup": r["matchup"],
                    "side": r["side"],
                    "player": f"slot {r['lineup_slot']} (id {r['batter_id']})",
                    "line": line,
                    "expected": r.get("expected_tb"),
                    "model_prob_over": r.get(key),
                })

        cols = ["market", "matchup", "side", "player", "line",
                "expected", "model_prob_over"]
        with path.open("w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

    @staticmethod
    def _write_xlsx(path: Path, payload: dict) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        wb.remove(wb.active)

        title_font = Font(bold=True, size=12, color="FFFFFF")
        title_fill = PatternFill("solid", fgColor="6B21A8")  # purple = sandbox
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="9333EA")

        for tab_name, key, columns in (
            ("Pitcher Ks", "pitcher_strikeouts", [
                "matchup", "side", "pitcher", "season_ks", "season_ip",
                "opp_team_k_per_9", "expected_ks",
                "over_4_5", "over_5_5", "over_6_5", "over_7_5",
            ]),
            ("Batter Hits", "batter_hits", [
                "matchup", "side", "lineup_slot", "batter_id",
                "season_avg", "season_ab", "opp_pitcher_baa",
                "expected_hits", "over_0_5", "over_1_5",
            ]),
            ("Batter Total Bases", "batter_total_bases", [
                "matchup", "side", "lineup_slot", "batter_id",
                "season_slg", "season_ab", "opp_pitcher_baa",
                "expected_tb", "over_1_5", "over_2_5", "over_3_5",
            ]),
        ):
            ws = wb.create_sheet(title=tab_name)
            ws.cell(row=1, column=1, value=f"EXPERIMENTAL — {tab_name} — {payload.get('target_date')}")
            ws.cell(row=1, column=1).font = title_font
            ws.cell(row=1, column=1).fill = title_fill
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

            for c, col in enumerate(columns, 1):
                cell = ws.cell(row=2, column=c, value=col)
                cell.font = header_font
                cell.fill = header_fill

            for r, row in enumerate(payload.get(key, []), 3):
                for c, col in enumerate(columns, 1):
                    ws.cell(row=r, column=c, value=row.get(col))

            for c, col in enumerate(columns, 1):
                ws.column_dimensions[ws.cell(row=2, column=c).column_letter].width = max(
                    12, len(col) + 2,
                )
            ws.freeze_panes = "A3"

        wb.save(path)


# Module-level scraper handle so static methods can reach back into the
# instance's batter-stats cache. Set during collect().
_MODULE_SCRAPER: MLBPlayerPropsScraper | None = None


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="MLB Player Props Spreadsheet (EXPERIMENTAL — sandbox only)",
    )
    parser.add_argument("--season", type=int, default=SEASON_DEFAULT)
    parser.add_argument("--date", type=str, default=None)
    parser.add_argument(
        "--output-dir", type=str, default=str(DEFAULT_OUTPUT_DIR),
    )
    args = parser.parse_args(argv)

    builder = PlayerPropsExperimental(
        season=args.season,
        target_date=args.date,
        output_dir=Path(args.output_dir),
    )
    global _MODULE_SCRAPER
    _MODULE_SCRAPER = builder.props_scraper

    print(f"\n[EXPERIMENTAL] MLB Player Props — target {builder.target_date}")
    print("This output is sandboxed; not surfaced on the website or daily card.\n")

    payload = builder.collect()
    written = builder.write(payload)

    print(f"\nWrote {len(written)} files to {builder.output_dir}:")
    for p in written:
        print(f"  - {p.relative_to(REPO_ROOT)}")

    counts = payload.get("counts", {})
    print(
        f"\nProjections: "
        f"{counts.get('pitcher_k_rows', 0)} pitcher Ks, "
        f"{counts.get('batter_hits_rows', 0)} batter hits, "
        f"{counts.get('batter_tb_rows', 0)} batter TB"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
