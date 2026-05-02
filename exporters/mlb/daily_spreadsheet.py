"""
MLB Daily Spreadsheet Exporter
==============================
Produces a 6-tab spreadsheet of game-results bets (Moneyline, Run Line,
Totals, First 5, First Inning, Team Totals) covering season-to-date
backfill plus projections for today's slate. Outputs land in
public/data/mlb/ so a Vercel-hosted frontend can serve them.

Usage:
    python -m exporters.mlb.daily_spreadsheet
    python -m exporters.mlb.daily_spreadsheet --date 2026-05-02
    python -m exporters.mlb.daily_spreadsheet --no-push
    python -m exporters.mlb.daily_spreadsheet --season 2026 --output-dir public/data/mlb
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None

import requests

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scrapers.mlb.mlb_game_scraper import MLBGameScraper, TEAM_MAP
from scrapers.mlb.mlb_odds_scraper import MLBOddsScraper
from exporters.mlb.projections import (
    ProjectionModel,
    prob_over,
    TOTAL_SD,
    TEAM_TOTAL_SD,
    F5_TOTAL_SD,
)
from exporters.mlb.kelly import kelly_advice, DEFAULT_DECIMAL_ODDS
from exporters.mlb.backtest import BacktestEngine


SEASON_DEFAULT = 2026
SEASON_OPENING_DAY = "{season}-03-20"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "public" / "data" / "mlb"
TOTAL_LINES = (8.5, 9.0, 9.5)
TEAM_TOTAL_LINES = (3.5, 4.5)


def _today_et() -> str:
    if ZoneInfo is not None:
        return datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d")
    return datetime.utcnow().strftime("%Y-%m-%d")


def _ou_label(actual: float, line: float) -> str:
    if actual > line:
        return "OVER"
    if actual < line:
        return "UNDER"
    return "PUSH"


def _best_total_kelly(mean: float, lines: tuple[float, ...], sd: float) -> dict:
    """Pick the (line, side) with the highest half-Kelly fraction at the default price."""
    best = None
    for line in lines:
        p_over = prob_over(line, mean, sd)
        for side, prob in (("OVER", p_over), ("UNDER", 1 - p_over)):
            adv = kelly_advice(prob)
            if best is None or adv["kelly_pct"] > best["kelly_pct"]:
                best = {**adv, "line": line, "side": side}
    return best


def _market_total_kelly(
    mean: float,
    sd: float,
    market_totals: list[dict],
) -> dict | None:
    """Best Kelly across whichever totals lines the book is actually offering."""
    best = None
    for offer in market_totals:
        line = offer.get("point")
        if line is None:
            continue
        p_over = prob_over(line, mean, sd)
        for side_key, side_label, prob in (
            ("over", "OVER", p_over),
            ("under", "UNDER", 1 - p_over),
        ):
            price = offer.get(side_key)
            if not price:
                continue
            adv = kelly_advice(prob, decimal_odds=price["decimal"])
            cand = {
                **adv,
                "line": line,
                "side": side_label,
                "market_decimal": price["decimal"],
                "market_american": price["american"],
                "book": price["book"],
            }
            if best is None or cand["kelly_pct"] > best["kelly_pct"]:
                best = cand
    return best


def _ml_market_price(market: dict, side: str) -> dict | None:
    return (market or {}).get(side)


def _rl_market_price(rl_offers: list[dict], side: str, point: float) -> dict | None:
    """Find the run-line price for (side, point); tolerate small numeric drift."""
    for o in rl_offers or []:
        if o["team"] == side and abs(o["point"] - point) < 0.01:
            return o
    return None


def fetch_slate(date: str) -> list[dict]:
    """Fetch every scheduled MLB game for a date (any status)."""
    url = (
        "https://statsapi.mlb.com/api/v1/schedule"
        f"?sportId=1&date={date}"
        "&fields=dates,date,games,gamePk,gameDate,status,detailedState,"
        "teams,away,home,team,id,name"
    )
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    slate = []
    for date_obj in data.get("dates", []):
        for game in date_obj.get("games", []):
            try:
                away_id = game["teams"]["away"]["team"]["id"]
                home_id = game["teams"]["home"]["team"]["id"]
            except KeyError:
                continue
            slate.append({
                "date": date_obj["date"],
                "game_pk": game.get("gamePk"),
                "game_time": game.get("gameDate"),
                "status": game.get("status", {}).get("detailedState"),
                "away_team": TEAM_MAP.get(away_id, str(away_id)),
                "home_team": TEAM_MAP.get(home_id, str(home_id)),
            })
    return slate


class DailySpreadsheet:
    """Builds and writes the daily MLB game-results spreadsheet."""

    BET_TABS = (
        "todays_card",
        "moneyline",
        "run_line",
        "totals",
        "first_5",
        "first_inning",
        "team_totals",
        "backtest",
    )

    def __init__(
        self,
        season: int = SEASON_DEFAULT,
        target_date: str | None = None,
        output_dir: Path | None = None,
        odds_api_key: str | None = None,
        skip_odds: bool = False,
    ):
        self.season = season
        self.target_date = target_date or _today_et()
        self.output_dir = Path(output_dir) if output_dir else DEFAULT_OUTPUT_DIR
        self.scraper = MLBGameScraper()
        self.odds_scraper = MLBOddsScraper(api_key=odds_api_key)
        self.skip_odds = skip_odds

    # --------- data assembly ------------------------------------------------

    def collect(self) -> dict:
        """Pull backfill + slate, compute projections, return structured data."""
        start = SEASON_OPENING_DAY.format(season=self.season)
        end = (
            datetime.strptime(self.target_date, "%Y-%m-%d") - timedelta(days=1)
        ).strftime("%Y-%m-%d")

        print(f"  Fetching backfill {start} -> {end}...")
        backfill = self.scraper.fetch_schedule(start, end)
        print(f"    {len(backfill)} completed games")

        print(f"  Fetching slate for {self.target_date}...")
        slate = fetch_slate(self.target_date)
        print(f"    {len(slate)} scheduled games")

        print("  Building projection model...")
        model = ProjectionModel(backfill)
        projections = model.project_slate(slate)

        if self.skip_odds:
            odds = {"fetched_at": None, "source": "skipped", "games": []}
        else:
            print("  Fetching market odds...")
            odds = self.odds_scraper.fetch()
            print(f"    {odds['source']} -> {len(odds['games'])} priced games")

        print("  Running backtest...")
        backtest = BacktestEngine(backfill).run()
        print(f"    {backtest['overall']['bets']} graded bets, "
              f"hit rate {backtest['overall']['hit_rate']}%, "
              f"units {backtest['overall']['units_pl']:+.2f}")

        tabs = {
            "moneyline": self._build_moneyline(backfill, projections, odds),
            "run_line": self._build_run_line(backfill, projections, odds),
            "totals": self._build_totals(backfill, projections, odds),
            "first_5": self._build_first_5(backfill, projections, odds),
            "first_inning": self._build_first_inning(backfill, projections, odds),
            "team_totals": self._build_team_totals(backfill, projections, odds),
        }
        tabs["todays_card"] = self._build_todays_card(tabs)
        tabs["backtest"] = self._build_backtest(backtest)

        return {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "season": self.season,
            "today": self.target_date,
            "odds_source": odds["source"],
            "odds_fetched_at": odds["fetched_at"],
            "counts": {
                "backfill_games": len(backfill),
                "slate_games": len(slate),
                "priced_games": len(odds["games"]),
            },
            "tabs": tabs,
            "odds": odds,
            "backtest": backtest,
        }

    # --------- per-tab builders --------------------------------------------

    @staticmethod
    def _build_moneyline(
        backfill: list[dict], projections: list[dict], odds: dict
    ) -> dict:
        proj_rows = []
        for p in projections:
            pick_side = "home" if p["ml_pick"] == p["home_team"] else "away"
            pick_prob = (
                p["home_win_prob"] if pick_side == "home" else p["away_win_prob"]
            )

            game_odds = MLBOddsScraper.find_game(odds, p["away_team"], p["home_team"])
            market = _ml_market_price((game_odds or {}).get("moneyline", {}), pick_side)

            if market:
                adv = kelly_advice(pick_prob, decimal_odds=market["decimal"])
                market_dec = market["decimal"]
                market_am = market["american"]
                book = market["book"]
            else:
                adv = kelly_advice(pick_prob)
                market_dec = None
                market_am = None
                book = None

            proj_rows.append({
                "date": p["date"],
                "away": p["away_team"],
                "home": p["home_team"],
                "away_runs_proj": p["away_runs_proj"],
                "home_runs_proj": p["home_runs_proj"],
                "away_win_prob": p["away_win_prob"],
                "home_win_prob": p["home_win_prob"],
                "ml_pick": p["ml_pick"],
                "model_prob": adv["model_prob"],
                "fair_odds_dec": adv["fair_odds_dec"],
                "market_odds_dec": market_dec,
                "market_odds_american": market_am,
                "book": book,
                "kelly_pct": adv["kelly_pct"],
                "kelly_advice": adv["kelly_advice"],
            })
        backfill_rows = [
            {
                "date": g["date"],
                "away": g["away_team"],
                "home": g["home_team"],
                "away_score": g["away_score"],
                "home_score": g["home_score"],
                "ml_winner": g["ml_winner"],
            }
            for g in sorted(backfill, key=lambda g: g["date"], reverse=True)
        ]
        return {
            "title": "Moneyline",
            "projection_columns": [
                "date", "away", "home", "away_runs_proj", "home_runs_proj",
                "away_win_prob", "home_win_prob", "ml_pick",
                "model_prob", "fair_odds_dec",
                "market_odds_dec", "market_odds_american", "book",
                "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "away", "home", "away_score", "home_score", "ml_winner",
            ],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_run_line(
        backfill: list[dict], projections: list[dict], odds: dict
    ) -> dict:
        proj_rows = []
        for p in projections:
            fav_side = "home" if p["rl_fav"] == p["home_team"] else "away"

            game_odds = MLBOddsScraper.find_game(odds, p["away_team"], p["home_team"])
            market = _rl_market_price(
                (game_odds or {}).get("run_line", []), fav_side, -1.5,
            )

            if market:
                adv = kelly_advice(p["rl_cover_prob"], decimal_odds=market["decimal"])
                market_dec = market["decimal"]
                market_am = market["american"]
                book = market["book"]
            else:
                adv = kelly_advice(p["rl_cover_prob"])
                market_dec = None
                market_am = None
                book = None

            proj_rows.append({
                "date": p["date"],
                "away": p["away_team"],
                "home": p["home_team"],
                "rl_fav": p["rl_fav"],
                "rl_margin_proj": p["rl_margin_proj"],
                "rl_cover_prob": p["rl_cover_prob"],
                "rl_fav_covers_1_5": "YES" if p["rl_fav_covers_1_5"] else "NO",
                "model_prob": adv["model_prob"],
                "fair_odds_dec": adv["fair_odds_dec"],
                "market_odds_dec": market_dec,
                "market_odds_american": market_am,
                "book": book,
                "kelly_pct": adv["kelly_pct"],
                "kelly_advice": adv["kelly_advice"],
            })
        backfill_rows = []
        for g in sorted(backfill, key=lambda g: g["date"], reverse=True):
            margin = g["away_score"] - g["home_score"]
            fav = g["away_team"] if margin > 0 else g["home_team"] if margin < 0 else "PUSH"
            backfill_rows.append({
                "date": g["date"],
                "away": g["away_team"],
                "home": g["home_team"],
                "away_score": g["away_score"],
                "home_score": g["home_score"],
                "margin": abs(margin),
                "favorite": fav,
                "fav_covers_1_5": "YES" if g["rl_favorite_covered"] else "NO",
            })
        return {
            "title": "Run Line",
            "projection_columns": [
                "date", "away", "home", "rl_fav", "rl_margin_proj",
                "rl_cover_prob", "rl_fav_covers_1_5",
                "model_prob", "fair_odds_dec",
                "market_odds_dec", "market_odds_american", "book",
                "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "away", "home", "away_score", "home_score",
                "margin", "favorite", "fav_covers_1_5",
            ],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_totals(
        backfill: list[dict], projections: list[dict], odds: dict
    ) -> dict:
        proj_rows = []
        for p in projections:
            row = {
                "date": p["date"],
                "away": p["away_team"],
                "home": p["home_team"],
                "away_runs_proj": p["away_runs_proj"],
                "home_runs_proj": p["home_runs_proj"],
                "total_proj": p["total_proj"],
            }
            for line in TOTAL_LINES:
                row[f"pick_{line}"] = "OVER" if p["total_proj"] > line else "UNDER"

            game_odds = MLBOddsScraper.find_game(odds, p["away_team"], p["home_team"])
            market_totals = (game_odds or {}).get("totals") or []
            best = _market_total_kelly(p["total_proj"], TOTAL_SD, market_totals)
            if best is None:
                best = _best_total_kelly(p["total_proj"], TOTAL_LINES, TOTAL_SD)
                row.update({
                    "kelly_line": f"{best['side']} {best['line']}",
                    "model_prob": best["model_prob"],
                    "fair_odds_dec": best["fair_odds_dec"],
                    "market_odds_dec": None,
                    "market_odds_american": None,
                    "book": None,
                    "kelly_pct": best["kelly_pct"],
                    "kelly_advice": best["kelly_advice"],
                })
            else:
                row.update({
                    "kelly_line": f"{best['side']} {best['line']}",
                    "model_prob": best["model_prob"],
                    "fair_odds_dec": best["fair_odds_dec"],
                    "market_odds_dec": best["market_decimal"],
                    "market_odds_american": best["market_american"],
                    "book": best["book"],
                    "kelly_pct": best["kelly_pct"],
                    "kelly_advice": best["kelly_advice"],
                })
            proj_rows.append(row)

        backfill_rows = []
        for g in sorted(backfill, key=lambda g: g["date"], reverse=True):
            row = {
                "date": g["date"],
                "away": g["away_team"],
                "home": g["home_team"],
                "total_runs": g["total"],
            }
            for line in TOTAL_LINES:
                row[f"result_{line}"] = _ou_label(g["total"], line)
            backfill_rows.append(row)

        return {
            "title": "Totals",
            "projection_columns": [
                "date", "away", "home", "away_runs_proj", "home_runs_proj", "total_proj",
            ] + [f"pick_{l}" for l in TOTAL_LINES] + [
                "kelly_line", "model_prob", "fair_odds_dec",
                "market_odds_dec", "market_odds_american", "book",
                "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "away", "home", "total_runs",
            ] + [f"result_{l}" for l in TOTAL_LINES],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_first_5(
        backfill: list[dict], projections: list[dict], odds: dict
    ) -> dict:
        proj_rows = []
        for p in projections:
            adv = kelly_advice(p["f5_win_prob"]) if p["f5_pick"] != "PUSH" \
                else {"model_prob": 0.5, "fair_odds_dec": 2.0,
                      "kelly_pct": 0.0, "kelly_advice": "PASS"}
            proj_rows.append({
                "date": p["date"],
                "away": p["away_team"],
                "home": p["home_team"],
                "f5_away_proj": p["f5_away_proj"],
                "f5_home_proj": p["f5_home_proj"],
                "f5_total_proj": p["f5_total_proj"],
                "f5_pick": p["f5_pick"],
                "f5_win_prob": p["f5_win_prob"],
                "model_prob": adv["model_prob"],
                "fair_odds_dec": adv["fair_odds_dec"],
                "kelly_pct": adv["kelly_pct"],
                "kelly_advice": adv["kelly_advice"],
            })
        backfill_rows = [
            {
                "date": g["date"],
                "away": g["away_team"],
                "home": g["home_team"],
                "f5_away": g["f5_away"],
                "f5_home": g["f5_home"],
                "f5_total": g["f5_away"] + g["f5_home"],
                "f5_winner": g["f5_winner"],
            }
            for g in sorted(backfill, key=lambda g: g["date"], reverse=True)
        ]
        return {
            "title": "First 5",
            "projection_columns": [
                "date", "away", "home", "f5_away_proj", "f5_home_proj",
                "f5_total_proj", "f5_pick", "f5_win_prob",
                "model_prob", "fair_odds_dec", "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "away", "home", "f5_away", "f5_home", "f5_total", "f5_winner",
            ],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_first_inning(
        backfill: list[dict], projections: list[dict], odds: dict
    ) -> dict:
        proj_rows = []
        for p in projections:
            pick_prob = p["nrfi_prob"] if p["nrfi_pick"] == "NRFI" else p["yrfi_prob"]
            adv = kelly_advice(pick_prob)
            proj_rows.append({
                "date": p["date"],
                "away": p["away_team"],
                "home": p["home_team"],
                "f1_total_proj": p["f1_total_proj"],
                "nrfi_prob": p["nrfi_prob"],
                "yrfi_prob": p["yrfi_prob"],
                "nrfi_pick": p["nrfi_pick"],
                "model_prob": adv["model_prob"],
                "fair_odds_dec": adv["fair_odds_dec"],
                "kelly_pct": adv["kelly_pct"],
                "kelly_advice": adv["kelly_advice"],
            })
        backfill_rows = [
            {
                "date": g["date"],
                "away": g["away_team"],
                "home": g["home_team"],
                "f1_away": g["f1_away"],
                "f1_home": g["f1_home"],
                "f1_total": g["f1_away"] + g["f1_home"],
                "nrfi_yrfi": "NRFI" if g["nrfi"] else "YRFI",
            }
            for g in sorted(backfill, key=lambda g: g["date"], reverse=True)
        ]
        return {
            "title": "First Inning",
            "projection_columns": [
                "date", "away", "home", "f1_total_proj",
                "nrfi_prob", "yrfi_prob", "nrfi_pick",
                "model_prob", "fair_odds_dec", "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "away", "home", "f1_away", "f1_home", "f1_total", "nrfi_yrfi",
            ],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_team_totals(
        backfill: list[dict], projections: list[dict], odds: dict
    ) -> dict:
        proj_rows = []
        for p in projections:
            for side, team, opp, runs in (
                ("AWAY", p["away_team"], p["home_team"], p["away_runs_proj"]),
                ("HOME", p["home_team"], p["away_team"], p["home_runs_proj"]),
            ):
                row = {
                    "date": p["date"],
                    "team": team,
                    "opponent": opp,
                    "side": side,
                    "team_total_proj": runs,
                }
                for line in TEAM_TOTAL_LINES:
                    row[f"pick_{line}"] = "OVER" if runs > line else "UNDER"
                best = _best_total_kelly(runs, TEAM_TOTAL_LINES, TEAM_TOTAL_SD)
                row.update({
                    "kelly_line": f"{best['side']} {best['line']}",
                    "model_prob": best["model_prob"],
                    "fair_odds_dec": best["fair_odds_dec"],
                    "kelly_pct": best["kelly_pct"],
                    "kelly_advice": best["kelly_advice"],
                })
                proj_rows.append(row)

        backfill_rows = []
        for g in sorted(backfill, key=lambda g: g["date"], reverse=True):
            for side, team, opp, runs in (
                ("AWAY", g["away_team"], g["home_team"], g["away_score"]),
                ("HOME", g["home_team"], g["away_team"], g["home_score"]),
            ):
                row = {
                    "date": g["date"],
                    "team": team,
                    "opponent": opp,
                    "side": side,
                    "runs": runs,
                }
                for line in TEAM_TOTAL_LINES:
                    row[f"result_{line}"] = _ou_label(runs, line)
                backfill_rows.append(row)

        return {
            "title": "Team Totals",
            "projection_columns": [
                "date", "team", "opponent", "side", "team_total_proj",
            ] + [f"pick_{l}" for l in TEAM_TOTAL_LINES] + [
                "kelly_line", "model_prob", "fair_odds_dec", "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "team", "opponent", "side", "runs",
            ] + [f"result_{l}" for l in TEAM_TOTAL_LINES],
            "projections": proj_rows,
            "backfill": backfill_rows,
        }

    @staticmethod
    def _build_todays_card(tabs: dict) -> dict:
        """Cross-tab summary: every projection row ranked by Kelly edge."""
        sources = (
            ("moneyline",   "ml_pick"),
            ("run_line",    "rl_fav"),
            ("totals",      "kelly_line"),
            ("first_5",     "f5_pick"),
            ("first_inning","nrfi_pick"),
            ("team_totals", "kelly_line"),
        )
        rows = []
        for tab_key, pick_key in sources:
            for r in tabs[tab_key]["projections"]:
                kelly = r.get("kelly_pct")
                if kelly is None:
                    continue
                # team_totals rows have team/opponent; others have away/home
                if "team" in r and "opponent" in r:
                    matchup = f"{r['team']} vs {r['opponent']}"
                else:
                    matchup = f"{r.get('away')}@{r.get('home')}"
                rows.append({
                    "date": r.get("date"),
                    "matchup": matchup,
                    "bet_type": tab_key,
                    "pick": r.get(pick_key),
                    "model_prob": r.get("model_prob"),
                    "fair_odds_dec": r.get("fair_odds_dec"),
                    "market_odds_dec": r.get("market_odds_dec"),
                    "market_odds_american": r.get("market_odds_american"),
                    "book": r.get("book"),
                    "kelly_pct": kelly,
                    "kelly_advice": r.get("kelly_advice"),
                })
        rows.sort(key=lambda r: r.get("kelly_pct") or 0, reverse=True)
        actionable = [r for r in rows if r.get("kelly_advice") and r["kelly_advice"] != "PASS"]
        passed = [r for r in rows if r.get("kelly_advice") == "PASS"]

        return {
            "title": "Today's Card",
            "projection_section_title":
                f"TODAY'S CARD — top picks by Kelly edge ({len(actionable)} plays)",
            "backfill_section_title":
                f"PASS list — model has no edge or negative EV ({len(passed)} skipped)",
            "projection_columns": [
                "date", "matchup", "bet_type", "pick", "model_prob",
                "fair_odds_dec", "market_odds_dec", "market_odds_american",
                "book", "kelly_pct", "kelly_advice",
            ],
            "backfill_columns": [
                "date", "matchup", "bet_type", "pick", "model_prob",
                "market_odds_dec", "kelly_pct",
            ],
            "projections": actionable,
            "backfill": passed,
        }

    @staticmethod
    def _build_backtest(backtest: dict) -> dict:
        overall = backtest["overall"]
        summary_rows = [{
            "scope": "OVERALL",
            "bet_type": "all",
            "bets": overall["bets"],
            "wins": overall["wins"],
            "losses": overall["losses"],
            "pushes": overall["pushes"],
            "hit_rate": overall["hit_rate"],
            "units_pl": overall["units_pl"],
            "roi_pct": overall["roi_pct"],
        }]
        for row in backtest["summary_by_bet_type"]:
            summary_rows.append({"scope": "BY TYPE", **row})

        return {
            "title": "Backtest",
            "projection_section_title":
                f"BACKTEST SUMMARY — "
                f"{backtest['first_date']} → {backtest['last_date']}",
            "backfill_section_title": "DAILY P&L (cumulative units, flat 1u @ -110)",
            "projection_columns": [
                "scope", "bet_type", "bets", "wins", "losses", "pushes",
                "hit_rate", "units_pl", "roi_pct",
            ],
            "backfill_columns": ["date", "daily_units", "cumulative_units"],
            "projections": summary_rows,
            "backfill": list(reversed(backtest["daily_pl"])),
        }

    # --------- writers -----------------------------------------------------

    def write_all(self, data: dict) -> list[Path]:
        """Write JSON, per-tab CSVs, and the multi-tab XLSX."""
        self.output_dir.mkdir(parents=True, exist_ok=True)
        written: list[Path] = []

        json_path = self.output_dir / "mlb_daily.json"
        json_path.write_text(json.dumps(data, indent=2, default=str))
        written.append(json_path)

        odds_payload = data.get("odds")
        if odds_payload is not None:
            odds_path = self.output_dir / "lines.json"
            odds_path.write_text(json.dumps(odds_payload, indent=2, default=str))
            written.append(odds_path)

        backtest_payload = data.get("backtest")
        if backtest_payload is not None:
            backtest_path = self.output_dir / "backtest.json"
            backtest_path.write_text(
                json.dumps(backtest_payload, indent=2, default=str)
            )
            written.append(backtest_path)

        for key in self.BET_TABS:
            tab = data["tabs"][key]
            csv_path = self.output_dir / f"{key}.csv"
            self._write_tab_csv(csv_path, tab)
            written.append(csv_path)

        xlsx_path = self.output_dir / "mlb_daily.xlsx"
        self._write_xlsx(xlsx_path, data)
        written.append(xlsx_path)

        return written

    @staticmethod
    def _write_tab_csv(path: Path, tab: dict) -> None:
        all_cols = ["section"] + sorted(
            set(tab["projection_columns"]) | set(tab["backfill_columns"]),
            key=lambda c: (
                tab["projection_columns"].index(c)
                if c in tab["projection_columns"]
                else len(tab["projection_columns"]) + tab["backfill_columns"].index(c)
            ),
        )
        with path.open("w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=all_cols, extrasaction="ignore")
            writer.writeheader()
            for row in tab["projections"]:
                writer.writerow({"section": "projection", **row})
            for row in tab["backfill"]:
                writer.writerow({"section": "backfill", **row})

    def _write_xlsx(self, path: Path, data: dict) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        wb.remove(wb.active)

        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2E75B6")
        backfill_title_fill = PatternFill("solid", fgColor="2E5984")

        for key in self.BET_TABS:
            tab = data["tabs"][key]
            ws = wb.create_sheet(title=tab["title"])

            proj_title = tab.get(
                "projection_section_title", f"PROJECTIONS — {data['today']}"
            )
            backfill_title = tab.get(
                "backfill_section_title",
                f"BACKFILL — {data['season']} season-to-date "
                f"({data['counts']['backfill_games']} games)",
            )

            ws.cell(row=1, column=1, value=proj_title)
            ws.cell(row=1, column=1).font = title_font
            ws.cell(row=1, column=1).fill = title_fill
            ws.merge_cells(
                start_row=1, start_column=1,
                end_row=1, end_column=max(len(tab["projection_columns"]), 1),
            )

            for c, col in enumerate(tab["projection_columns"], 1):
                cell = ws.cell(row=2, column=c, value=col)
                cell.font = header_font
                cell.fill = header_fill
            for r, row in enumerate(tab["projections"], 3):
                for c, col in enumerate(tab["projection_columns"], 1):
                    ws.cell(row=r, column=c, value=row.get(col))

            backfill_start = 3 + len(tab["projections"]) + 1
            ws.cell(row=backfill_start, column=1, value=backfill_title)
            ws.cell(row=backfill_start, column=1).font = title_font
            ws.cell(row=backfill_start, column=1).fill = backfill_title_fill
            ws.merge_cells(
                start_row=backfill_start, start_column=1,
                end_row=backfill_start, end_column=max(len(tab["backfill_columns"]), 1),
            )

            for c, col in enumerate(tab["backfill_columns"], 1):
                cell = ws.cell(row=backfill_start + 1, column=c, value=col)
                cell.font = header_font
                cell.fill = header_fill
            for r, row in enumerate(tab["backfill"], backfill_start + 2):
                for c, col in enumerate(tab["backfill_columns"], 1):
                    ws.cell(row=r, column=c, value=row.get(col))

            for col_idx, col in enumerate(
                set(tab["projection_columns"]) | set(tab["backfill_columns"]), 1
            ):
                ws.column_dimensions[
                    ws.cell(row=2, column=col_idx).column_letter
                ].width = max(12, len(col) + 2)

            ws.freeze_panes = "A3"

        wb.save(path)

    # --------- git ---------------------------------------------------------

    def commit_and_push(
        self,
        files: Iterable[Path],
        branch: str | None = None,
    ) -> bool:
        """Stage, commit, and push the written files. Returns True on success."""
        rel = [str(p.relative_to(REPO_ROOT)) for p in files]

        try:
            subprocess.run(
                ["git", "-C", str(REPO_ROOT), "add", *rel],
                check=True, capture_output=True, text=True,
            )
            status = subprocess.run(
                ["git", "-C", str(REPO_ROOT), "status", "--porcelain", *rel],
                check=True, capture_output=True, text=True,
            )
            if not status.stdout.strip():
                print("  No changes to commit.")
                return True

            msg = f"Daily MLB spreadsheet — {self.target_date}"
            subprocess.run(
                ["git", "-C", str(REPO_ROOT), "commit", "-m", msg],
                check=True, capture_output=True, text=True,
            )
            print(f"  Committed: {msg}")

            push_cmd = ["git", "-C", str(REPO_ROOT), "push"]
            if branch:
                push_cmd += ["-u", "origin", branch]
            subprocess.run(push_cmd, check=True, capture_output=True, text=True)
            print("  Pushed.")
            return True
        except subprocess.CalledProcessError as e:
            print(f"  git failed: {e.stderr or e.stdout}")
            return False


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="MLB Daily Spreadsheet Exporter")
    parser.add_argument("--season", type=int, default=SEASON_DEFAULT)
    parser.add_argument(
        "--date", type=str, default=None,
        help="Target date in YYYY-MM-DD (defaults to today, ET)",
    )
    parser.add_argument(
        "--output-dir", type=str, default=str(DEFAULT_OUTPUT_DIR),
        help="Directory to write outputs into",
    )
    parser.add_argument(
        "--push", action="store_true", default=False,
        help="After writing, git add/commit/push the new files",
    )
    parser.add_argument(
        "--branch", type=str, default=None,
        help="Branch to push to (only with --push)",
    )
    parser.add_argument(
        "--odds-api-key", type=str, default=None,
        help="The Odds API key (overrides ODDS_API_KEY env var)",
    )
    parser.add_argument(
        "--no-odds", action="store_true", default=False,
        help="Skip the odds fetch entirely; use -110 default for all Kelly sizing",
    )
    args = parser.parse_args(argv)

    spreadsheet = DailySpreadsheet(
        season=args.season,
        target_date=args.date,
        output_dir=Path(args.output_dir),
        odds_api_key=args.odds_api_key,
        skip_odds=args.no_odds,
    )

    print(f"MLB Daily Spreadsheet — target {spreadsheet.target_date}")
    data = spreadsheet.collect()
    written = spreadsheet.write_all(data)

    print(f"Wrote {len(written)} files to {spreadsheet.output_dir}:")
    for p in written:
        print(f"  - {p.relative_to(REPO_ROOT)}")

    if args.push:
        ok = spreadsheet.commit_and_push(written, branch=args.branch)
        return 0 if ok else 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
